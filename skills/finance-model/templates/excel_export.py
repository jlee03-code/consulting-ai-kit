"""
BCG finance model -> Excel (.xlsx) exporter.
Sources rendered as clickable hyperlinks in a dedicated Sources sheet.
Requires: pip install openpyxl
Usage: python3 excel_export.py
"""

from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@dataclass
class Source:
    name: str
    url: str  # empty string -> flagged as unverifiable


@dataclass
class FinanceOutput:
    headline: str
    assumptions: dict       # {label: (value, cell_type)}  cell_type: '입력'/'수식'/'연결'
    projections: dict       # {metric: {year: value}}
    years: list             # e.g. ['2024A', '2025F', ..., 'CAGR']
    valuation: dict         # {metric: {'Base': val, 'Bull': val, 'Bear': val}}
    sensitivity: dict       # {wacc_float: {tgr_float: ev_value}}
    wacc_range: list
    tgr_range: list
    bcg_opinion: str
    so_what: str
    sources: list           # list of Source


# Palette
BCG_GREEN = '006A4E'
INPUT_BLUE = 'DBEAFE'
FORMULA_GRAY = 'F3F4F6'
LINK_TEAL = 'CCFBF1'

_THIN = Side(style='thin', color='CCCCCC')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr_fill(color: str) -> PatternFill:
    return PatternFill(fill_type='solid', fgColor=color)


def _set_header_row(ws, row: int, values: list, bg: str = BCG_GREEN) -> None:
    fill = _hdr_fill(bg)
    txt_color = 'FFFFFF' if bg == BCG_GREEN else '111111'
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color=txt_color, size=10)
        c.fill = fill
        c.border = _BORDER
        c.alignment = Alignment(horizontal='center', wrap_text=True)


def _add_sources_sheet(wb: Workbook, sources: list) -> None:
    ws = wb.create_sheet('Sources')
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 60
    _set_header_row(ws, 1, ['Source', 'Link'])

    for i, src in enumerate(sources, 2):
        ws.cell(row=i, column=1, value=src.name).border = _BORDER
        link_cell = ws.cell(row=i, column=2)
        if src.url:
            link_cell.value = src.name
            link_cell.hyperlink = src.url
            link_cell.font = Font(color='0563C1', underline='single', size=10)
        else:
            link_cell.value = '출처 불명확 — 추가 검증 필요'
            link_cell.font = Font(color='DC2626', size=10)
        link_cell.border = _BORDER


def _add_assumptions_sheet(wb: Workbook, output: FinanceOutput) -> None:
    ws = wb.create_sheet('Assumptions')
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    _set_header_row(ws, 1, ['Assumption', 'Value', 'Type'])

    type_fill = {
        '입력': _hdr_fill(INPUT_BLUE),
        '수식': _hdr_fill(FORMULA_GRAY),
        '연결': _hdr_fill(LINK_TEAL),
    }

    for i, (label, (value, cell_type)) in enumerate(output.assumptions.items(), 2):
        ws.cell(row=i, column=1, value=label).border = _BORDER
        val_c = ws.cell(row=i, column=2, value=value)
        val_c.border = _BORDER
        val_c.fill = type_fill.get(cell_type, type_fill['입력'])
        val_c.alignment = Alignment(horizontal='right')
        type_c = ws.cell(row=i, column=3, value=f'[{cell_type}]')
        type_c.border = _BORDER
        type_c.alignment = Alignment(horizontal='center')
        type_c.font = Font(size=9, color='555555')


def _add_projections_sheet(wb: Workbook, output: FinanceOutput) -> None:
    ws = wb.create_sheet('Projections')
    ws.column_dimensions['A'].width = 24

    headers = ['(단위: $M)'] + output.years
    _set_header_row(ws, 1, headers)
    for col in range(2, len(output.years) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 13

    for r, (metric, year_values) in enumerate(output.projections.items(), 2):
        ws.cell(row=r, column=1, value=metric).border = _BORDER
        for c, year in enumerate(output.years, 2):
            cell = ws.cell(row=r, column=c, value=year_values.get(year, ''))
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal='right')
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.0'


def _add_valuation_sheet(wb: Workbook, output: FinanceOutput) -> None:
    ws = wb.create_sheet('Valuation')
    ws.column_dimensions['A'].width = 30
    for col in 'BCD':
        ws.column_dimensions[col].width = 16

    _set_header_row(ws, 1, ['(단위: $M)', 'Base', 'Bull', 'Bear'])

    for r, (metric, scenarios) in enumerate(output.valuation.items(), 2):
        ws.cell(row=r, column=1, value=metric).border = _BORDER
        for c, scenario in enumerate(['Base', 'Bull', 'Bear'], 2):
            cell = ws.cell(row=r, column=c, value=scenarios.get(scenario, ''))
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal='right')
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'


def _add_sensitivity_sheet(wb: Workbook, output: FinanceOutput) -> None:
    ws = wb.create_sheet('Sensitivity')
    ws.column_dimensions['A'].width = 24
    for col in range(2, len(output.tgr_range) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 13

    col_headers = ['EV ($M)  WACC \\ TGR'] + [f'{t:.1%}' for t in output.tgr_range]
    _set_header_row(ws, 1, col_headers)

    for r, wacc in enumerate(output.wacc_range, 2):
        ws.cell(row=r, column=1, value=f'{wacc:.1%}').border = _BORDER
        row_vals = output.sensitivity.get(wacc, {})
        for c, tgr in enumerate(output.tgr_range, 2):
            cell = ws.cell(row=r, column=c, value=row_vals.get(tgr, ''))
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal='right')
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'


def _add_opinion_sheet(wb: Workbook, output: FinanceOutput) -> None:
    ws = wb.create_sheet('BCG 의견')
    ws.column_dimensions['A'].width = 80
    _set_header_row(ws, 1, [output.headline])

    ws.cell(row=3, column=1, value='[BCG 의견]').font = Font(bold=True, color=BCG_GREEN, size=11)
    body = ws.cell(row=4, column=1, value=output.bcg_opinion)
    body.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 60

    ws.cell(row=6, column=1, value=f'특히, {output.so_what}').font = Font(bold=True, size=11)


def export_to_excel(output: FinanceOutput, filepath: str) -> str:
    """Write FinanceOutput to a formatted .xlsx and return the filepath."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _add_assumptions_sheet(wb, output)
    _add_projections_sheet(wb, output)
    _add_valuation_sheet(wb, output)
    _add_sensitivity_sheet(wb, output)
    _add_opinion_sheet(wb, output)
    _add_sources_sheet(wb, output.sources)

    wb.save(filepath)
    return filepath


if __name__ == '__main__':
    result = export_to_excel(
        FinanceOutput(
            headline='Subject DCF | EV ~$147B, implied price $52.00/주',
            assumptions={
                '매출 성장률 Y1': ('25%', '입력'),
                '매출 성장률 Y2': ('22%', '입력'),
                'EBITDA 마진': ('16%', '입력'),
                'WACC': ('10%', '입력'),
                'Terminal growth rate': ('2.5%', '입력'),
                'Tax rate': ('21%', '입력'),
                'Net debt ($M)': (-19000, '입력'),
                'Shares outstanding (M)': (3190, '입력'),
                'Enterprise Value ($M)': ('=PV FCF + PV TV', '수식'),
                'Equity Value ($M)': ('=EV − Net debt', '수식'),
            },
            projections={
                'Revenue':       {'2025F': 121000, '2026F': 147620, '2027F': 174192, '2028F': 200320, '2029F': 224359, 'CAGR': '16.7%'},
                'EBITDA':        {'2025F': 19360,  '2026F': 23619,  '2027F': 27871,  '2028F': 32051,  '2029F': 35897,  'CAGR': '16.7%'},
                'NOPAT':         {'2025F': 12427,  '2026F': 15161,  '2027F': 17890,  '2028F': 20573,  '2029F': 23042,  'CAGR': '16.7%'},
                'Unlevered FCF': {'2025F': 10007,  '2026F': 12353,  '2027F': 14819,  '2028F': 17263,  '2029F': 19596,  'CAGR': '18.3%'},
            },
            years=['2025F', '2026F', '2027F', '2028F', '2029F', 'CAGR'],
            valuation={
                'PV FCF ($M)':         {'Base': 56200,  'Bull': 64600,  'Bear': 48100},
                'Terminal Value ($M)':  {'Base': 107800, 'Bull': 128900, 'Bear': 89400},
                'EV ($M)':             {'Base': 146900, 'Bull': 182000, 'Bear': 121800},
                'Net Debt ($M)':       {'Base': -19000, 'Bull': -19000, 'Bear': -19000},
                'Equity Value ($M)':   {'Base': 165900, 'Bull': 201000, 'Bear': 140800},
                'Implied price/share': {'Base': 52.00,  'Bull': 63.00,  'Bear': 44.10},
            },
            sensitivity={
                0.08: {0.015: 195000, 0.020: 210000, 0.025: 228000, 0.030: 250000, 0.035: 278000},
                0.09: {0.015: 175000, 0.020: 188000, 0.025: 203000, 0.030: 221000, 0.035: 243000},
                0.10: {0.015: 158000, 0.020: 169000, 0.025: 182000, 0.030: 197000, 0.035: 215000},
                0.11: {0.015: 143000, 0.020: 153000, 0.025: 164000, 0.030: 177000, 0.035: 192000},
                0.12: {0.015: 130000, 0.020: 139000, 0.025: 148000, 0.030: 160000, 0.035: 173000},
            },
            wacc_range=[0.08, 0.09, 0.10, 0.11, 0.12],
            tgr_range=[0.015, 0.020, 0.025, 0.030, 0.035],
            bcg_opinion=(
                'TV가 EV의 73% 수준으로 terminal growth 가정에 민감. '
                'WACC 1pp 상승 시 EV ~12% 하락. '
                '성장률 가정은 업종 평균 대비 상단 — 추가 검증 필요.'
            ),
            so_what='TV 비중 70%+ 확인 → 따라서 terminal growth 가정 재검토 및 업종 비교군 추가 분석 권고',
            sources=[
                Source('회사 공개 연간보고서 2024', 'https://ir.example.com/annual-report-2024'),
                Source('Bloomberg Terminal — EV/EBITDA 업종 평균', ''),
                Source('BCG 추정', ''),
            ],
        ),
        '/tmp/bcg_finance_model.xlsx',
    )
    print(f'Saved: {result}')
    print('Sources 탭에서 링크를 클릭하면 브라우저에서 열립니다.')
